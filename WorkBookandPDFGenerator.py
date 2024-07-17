from msilib.schema import Font
from configparser import ConfigParser
import os
import openpyxl
import pandas as pd
import datetime
from datetime import timedelta, datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from SeparatePicklistPDFGenerator import generate_separate_picklist_pdf
from PicklistMerge import combine_pdf_files
from MonthlyForecastPDFGenerator import generate_monthly_forecast_pdf
from WeeklyForecastPDFGenerator import generate_weekly_forecast_pdf
from Configparser_ini_creator import updateConfigForPathLocations
from RemoveLastSavedPDFFiles import RemoveLastSavedPDFFiles
from servicesPicklistPDFGenerator import generate_services_picklist_pdf

#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------#

# Section - 1: Generating the workbook from HSE files and creating work order sheets Account wise for all weeks, and Weekly, Monthly Forecast files and Picklist sheet for 4 weeks. #

# Update the configuration with paths relative to the current script location
updateConfigForPathLocations()

def get_user_date():
    user_input = input("Enter a date in 'yyyy, m, d' format or type 'p' to use the current date: ")
    if user_input.lower() == 'p':
        return datetime.now()
    else:
        try:
            year, month, day = map(int, user_input.split(', '))
            return datetime(year, month, day)
        except ValueError:
            print("Sorry, that is in the incorrect format. Try again!")
            return get_user_date()

# Get the date from the user or use the current date
current_date = get_user_date()
print("The current date is:", current_date)

# Initialize the ConfigParser
config = ConfigParser()

# Read the config.ini file
config.read('config.ini')

# print(config.sections())

# Access the paths
hse_directory_path = config['PATHS']['hse_directory_path'] 
logo_path = config['PATHS']['logo_path']
output_directory_path = config['PATHS']['output_directory_path']
separate_sheets_output_directory_path = config['PATHS']['separate_sheets_output_directory_path']
formatted_current_date = current_date.strftime('%d_%m_%Y')
excel_file_path = config['PATHS']['excel_file_path'].replace('{current_date}', formatted_current_date)
full_picklist_pdf_output_directory_path = config['PATHS']['full_picklist_pdf_output_directory_path']
forecast_output_directory_path = config['PATHS']['forecast_output_directory_path']
separate_pdf_output_directory_path = config['PATHS']['separate_pdf_output_directory_path']
required_prices_path = config['PATHS']['required_prices_path']
services_picklist_output_directory_path = config['PATHS']['services_picklist_output_directory_path']

# print(required_prices_path)

RemoveLastSavedPDFFiles(separate_sheets_output_directory_path, services_picklist_output_directory_path, forecast_output_directory_path)

def createExcelWorkbookfromHSEFiles():

    # Calculate the start of the current month
    start_of_current_month = current_date.replace(day=1)

    # Define the order of the months starting from the current month
    months_order = pd.date_range(start=start_of_current_month, periods=12, freq='MS').strftime('%B %Y')

    # Calculate the Monday of the current week
    current_week_monday = current_date - timedelta(days=current_date.weekday())

    # Adjust current_week_monday to have 00:00 as time
    current_week_monday = datetime(current_week_monday.year, current_week_monday.month, current_week_monday.day)

    # Define the mapping of file prefixes to customer sheet names
    prefix_to_sheet = {
        "EMVX": "BAM002",
        "WPCX": "BAM003",
        "ATTX": "BAM004",
        "SPEX": "BAM005",
        "LPWX": "BAM007",
        "JHPX": "BAM008",
        "BHLX": "BAM009",
        "LDLX": "BAM011",
        "PWRX": "BAM018"
    }

    # Read the 'RequiredPrices' Excel sheet into a DataFrame
    required_prices_df = pd.read_excel(required_prices_path, sheet_name='RequiredPrices')

    def read_hse_file(file_path):
        # Updated column widths based on user input
        col_specs = [
            (0, 10),  # stock code
            (11, 13),  # issue
            (14, 22),  # required date
            (23, 31),  # required quantities
            (32, 42),  # order reference
            (43, 48),  # location
            (55, 75),  # message
            (76, 91),  # last delivery note
            (92, 100)  # last delivery date
        ]
        df = pd.read_fwf(file_path, colspecs=col_specs, header=None)
        df.columns = [
            'Stock Code', 'Issue', 'Required Date', 'Required Quantities',
            'Order Reference', 'Location', 'Message', 'Last Delivery Note',
            'Last Delivery Date'
        ]

        # Convert 'Required Date' to datetime format
        df['Required Date'] = pd.to_datetime(df['Required Date'], format='%Y%m%d', errors='coerce')

        # Convert 'Required Quantities' to string and remove leading zeros for comparison
        df['Required Quantities'] = df['Required Quantities'].astype(str).str.lstrip('0')

        # Filter out rows with 'Required Quantities' equal to "0" or empty after stripping leading zeros
        df = df[(df['Required Quantities'] != "0") & (df['Required Quantities'] != "")]

        # Filter to include rows with 'Required Date' from the Monday of the current week onwards
        # Use .copy() to avoid SettingWithCopyWarning
        if prefix == "WPCX":
            # Include all rows for "WPCX" prefix
            df_filtered = df.copy()
        else:
            # Filter to include rows with 'Required Date' from the Monday of the current week onwards
            df_filtered = df[df['Required Date'] >= current_week_monday].copy()

        # Exclude rows with specific stock codes
        excluded_stock_codes = ['400/U4238', '400/U4239', '331/62858']
        df_filtered = df_filtered[~df_filtered['Stock Code'].isin(excluded_stock_codes)]

        # Generating 'Arrears' for every BAM/Division!!
        df_filtered['Week'] = df_filtered['Required Date'].apply(lambda x: "Arrears" if x < current_week_monday else ((x - current_week_monday).days // 7) + 1)
        df_filtered['Week'] = df_filtered['Week'].apply(lambda x: f"Week - {x}" if isinstance(x, int) else x)

        # Format the 'Month' column to include the year
        df_filtered['Month'] = df_filtered['Required Date'].dt.strftime('%B %Y')

        # Convert 'Last Delivery Date' to the desired format
        df_filtered['Last Delivery Date'] = pd.to_datetime(df_filtered['Last Delivery Date'], format='%Y%m%d', errors='coerce').dt.strftime('%d/%m/%Y')

        # Convert 'Required Date' back to string format if needed
        df_filtered['Required Date'] = df_filtered['Required Date'].dt.strftime('%d/%m/%Y')

        # Drop duplicate rows based on 'Stock Code', 'Issue', 'Required Date', and 'Required Quantities'
        df_filtered = df_filtered.drop_duplicates(subset=['Stock Code', 'Issue', 'Required Date', 'Required Quantities'])

        return df_filtered

# Format the current date as dd_mm_YYYY
    formatted_date = current_date.strftime('%d_%m_%Y')

    # Create the output filename with the current date
    output_filename = f'SalesOrder_{formatted_date}.xlsx'
    output_filepath = os.path.join(output_directory_path, output_filename)

    # Create a Pandas Excel writer using XlsxWriter as the engine
    with pd.ExcelWriter(output_filepath, engine='xlsxwriter') as writer:
        # Dictionary to keep track of DataFrames for each sheet to avoid duplicates
        sheet_to_df = {}

        # Loop through the files in the directory
        for prefix, sheet_name in prefix_to_sheet.items():
            # Get all the files with the current prefix
            prefix_files = sorted([f for f in os.listdir(hse_directory_path) if f.startswith(prefix) and f.endswith('.hse')], key=lambda x: os.path.getmtime(os.path.join(hse_directory_path, x)))

            # Check if any files with the 'PWRX' prefix exist
            if prefix == 'PWRX' and not prefix_files:
                print(f"No HSE files found with prefix 'PWRX'. Skipping 'BAM018' sheet generation.")
                continue

            # Process the files in order of creation/modification
            for filename in prefix_files:
                # Read the .hse file into a DataFrame
                df = read_hse_file(os.path.join(hse_directory_path, filename))

                # If the sheet already has a DataFrame, update it based on the conditions
            if sheet_name in sheet_to_df:
                existing_df = sheet_to_df[sheet_name]

                # Convert 'Issue' column to string data type in both DataFrames
                existing_df['Issue'] = existing_df['Issue'].astype(str)
                df['Issue'] = df['Issue'].astype(str)

                # Merge the existing DataFrame with the new DataFrame based on 'Stock Code', 'Issue', and 'Required Date'
                merged_df = pd.merge(existing_df, df, on=['Stock Code', 'Issue', 'Required Date'], how='outer', suffixes=('', '_new'))

                # Update 'Required Quantities' with the new values where available
                merged_df['Required Quantities'] = merged_df.apply(lambda x: x['Required Quantities_new'] if pd.notnull(x['Required Quantities_new']) else x['Required Quantities'], axis=1)

                # Drop the temporary '_new' columns
                merged_df = merged_df.drop(columns=[col for col in merged_df.columns if col.endswith('_new')])

                # Remove rows where 'Required Quantities' is 0
                merged_df = merged_df[merged_df['Required Quantities'] != 0]

                sheet_to_df[sheet_name] = merged_df
            else:
                sheet_to_df[sheet_name] = df

        # Process each sheet's DataFrame
        for sheet_name, df in sheet_to_df.items():
            # Drop duplicate rows based on 'Stock Code' and 'Required Date'
            df = df.drop_duplicates(subset=['Stock Code', 'Issue', 'Required Date', 'Required Quantities'])

            # Merge the DataFrame with the 'RequiredPrices' DataFrame based on 'Stock Code'
            merged_df = pd.merge(df, required_prices_df, on='Stock Code', how='left')

            # Ensure 'Unit Price' and 'Required Quantities' are numeric
            merged_df['Unit Price'] = pd.to_numeric(merged_df['Unit Price'], errors='coerce')
            merged_df['Required Quantities'] = pd.to_numeric(merged_df['Required Quantities'], errors='coerce')

            # Now perform the multiplication
            merged_df['Sale Price'] = merged_df['Unit Price'] * merged_df['Required Quantities']

            # Drop duplicates again after the merge to ensure no duplicates based on 'Stock Code', 'Issue', and 'Required Date'
            merged_df = merged_df.drop_duplicates(subset=['Stock Code', 'Issue', 'Required Date'])

            # Write the merged DataFrame to the corresponding sheet in the Excel workbook
            merged_df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Get the xlsxwriter workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Get the dimensions of the dataframe
            (max_row, max_col) = merged_df.shape

            # Create a table with the dataframe data
            column_settings = [{'header': column} for column in merged_df.columns]
            worksheet.add_table(0, 0, max_row, max_col - 1, {'columns': column_settings})

            # Write the column headers with the defined format
            for col_num, value in enumerate(merged_df.columns.values):
                worksheet.write(0, col_num, value)

    # The Excel file 'output_with_prices.xlsx' is saved at this point with multiple sheets formatted as tables

    # Load the workbook to read existing sheets (customer accounts)
    wb = load_workbook(output_filepath)
    sheet_names = wb.sheetnames

    # Initialize an empty DataFrame for the summary
    weekly_df = pd.DataFrame()

    # Loop through each customer account sheet
    for sheet_name in sheet_names:
        if sheet_name in prefix_to_sheet.values():
            df = pd.read_excel(output_filepath, sheet_name=sheet_name)

            # Eliminate duplicate rows based on 'Stock Code' and 'Required Date'
            df = df.drop_duplicates(subset=['Stock Code', 'Issue', 'Required Date', 'Required Quantities'])

            # Clear the existing data in the sheet
            ws = wb[sheet_name]
            ws.delete_rows(2, ws.max_row + 1)
            
            # Write the deduplicated DataFrame back to the sheet
            for r_idx, row in df.iterrows():
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx + 2, column=c_idx, value=value)
            
            # Ensure 'Sale Price' is numeric for aggregation
            df['Sale Price'] = pd.to_numeric(df['Sale Price'], errors='coerce')

            # Filter data to include only 'Week - 1' onwards
            df = df[df['Week'] != 'Arrears']
            
            # Aggregate 'Sale Price' data by 'Week' and 'Customer Account'
            df = df.drop_duplicates(subset=['Stock Code', 'Required Date'])
            agg_data = df.groupby('Week').agg({'Sale Price': 'sum'}).reset_index()
            agg_data['Account'] = sheet_name  # Add the customer account column
            
            # Append the aggregated data to the Weekly Forecast DataFrame
            weekly_df = pd.concat([weekly_df, agg_data])

    # Pivot the Weekly Forecast DataFrame to get 'Customer Account' as rows and each week as columns
    pivot_df = pd.pivot_table(weekly_df, values='Sale Price', 
                            index=['Account'], columns=['Week'], aggfunc=sum, fill_value=0)

    # Sort the week columns in ascending order
    week_columns = sorted(pivot_df.columns, key=lambda x: int(x.split(' - ')[1]))
    pivot_df = pivot_df[week_columns]

    # Filter the columns to include only 'Week - 1' to 'Week - 52'
    week_columns = [f'Week - {i}' for i in range(1, 53)]
    pivot_df = pivot_df.reindex(columns=week_columns, fill_value=0)

    # Reset the index to bring 'Customer Account' back as a column
    pivot_df.reset_index(inplace=True)

    # Write the Weekly Forecast DataFrame to a new sheet in the same workbook
    with pd.ExcelWriter(output_filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        pivot_df.to_excel(writer, sheet_name='Weekly Forecast', index=False)

    print(f"Weekly Forecast written to {output_filepath} in the 'Weekly Forecast' sheet.")

    # Load the workbook and the specific sheet ('Weekly')
    wb = load_workbook(output_filepath)
    ws = wb['Weekly Forecast']  # Assuming 'Weekly' is the sheet name you want to add a table to

    # Determine the range of the table
    # Assuming data starts in the first row and first column
    # Find the last row with data
    last_row = ws.max_row
    # Find the last column with data
    last_col = ws.max_column

    # Convert the last column number to its corresponding letter
    last_col_letter = openpyxl.utils.get_column_letter(last_col)

    # Define the table range (e.g., "A1:D10")
    table_range = f"A1:{last_col_letter}{last_row}"

    # Create a table
    table = Table(displayName="WeeklySales", ref=table_range)

    # Add a default table style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style

    # Add the table to the worksheet
    ws.add_table(table)

    # Save the workbook
    wb.save(output_filepath)

    print(f"Table added to the 'Weekly' sheet in {output_filepath}")

    # Create a summary DataFrame for sales aggregates per month per customer account

    # Load the workbook to read existing sheets (customer accounts)
    wb = load_workbook(output_filepath)
    sheet_names = wb.sheetnames

    monthly_summary_data = []

    # Loop through each customer account sheet
    for sheet_name in prefix_to_sheet.values():
        if sheet_name in wb.sheetnames:
            df = pd.read_excel(output_filepath, sheet_name=sheet_name)

            # Eliminate duplicate rows based on 'Stock Code' and 'Required Date'
            df = df.drop_duplicates(subset=['Stock Code', 'Required Date'])
            df['Sale Price'] = pd.to_numeric(df['Sale Price'], errors='coerce')

            # Filter data to include only 'Week - 1' onwards
            df = df[df['Week'] != 'Arrears']

            # Aggregate 'Sale Price' data by 'Month'
            agg_data = df.groupby('Month').agg({'Sale Price': 'sum'}).reset_index()
            agg_data['Account'] = sheet_name  # Add the account column
            
            # Append the aggregated data to the monthly summary list
            monthly_summary_data.append(agg_data)
        else:
            print(f"Skipping sheet {sheet_name} because it was not found in the workbook.")

    # Concatenate all the monthly summary DataFrames
    monthly_summary_df = pd.concat(monthly_summary_data, ignore_index=True)

    # Reindex the DataFrame according to the defined order of months
    monthly_summary_df['Month'] = pd.Categorical(monthly_summary_df['Month'], categories=months_order, ordered=True)
    monthly_summary_df.sort_values('Month', inplace=True)

    # Pivot the monthly summary DataFrame to get 'Account' as rows and 'Month' as columns
    monthly_pivot_df = pd.pivot_table(monthly_summary_df, values='Sale Price', 
                                    index=['Account'], columns=['Month'], aggfunc=sum, fill_value=0)

    # Reset index to turn the index into a column
    monthly_pivot_df.reset_index(inplace=True)

    # Write the monthly summary pivot table to the 'Monthly Forecast' sheet
    with pd.ExcelWriter(output_filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        monthly_pivot_df.to_excel(writer, sheet_name='Monthly Forecast', index=False)

    wb = load_workbook(output_filepath)
    ws = wb['Monthly Forecast']

    # Determine the range of the table
    # Assuming data starts in the first row and first column
    # Find the last row with data
    last_row = ws.max_row
    # Find the last column with data
    last_col = ws.max_column

    # Convert the last column number to its corresponding letter
    last_col_letter = openpyxl.utils.get_column_letter(last_col)

    # Define the table range (e.g., "A1:D10")
    table_range = f"A1:{last_col_letter}{last_row}"

    # Create a table
    table = Table(displayName="MonthlySales", ref=table_range)    
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style

    # Add the table to the worksheet
    ws.add_table(table)

    # Save the workbook
    wb.save(output_filepath)

    print(f"Table added to the 'Monthly Forecast' sheet in {output_filepath}")

    wb = load_workbook(output_filepath)

    # Define the desired column widths for each sheet
    customer_sheet_width = 18
    weekly_forecast_width = 15
    monthly_forecast_width = 18

    # Define a custom style for the total cells
    total_style = Font(bold=True)
    total_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Function to adjust column widths for a sheet
    def adjust_column_widths(sheet_name, column_width):
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for col in ws.columns:
                ws.column_dimensions[get_column_letter(col[0].column)].width = column_width

    # Function to add total row and column for 'Weekly Forecast' and 'Monthly Forecast' sheets
    def add_totals(sheet_name):
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            max_row = ws.max_row
            max_col = ws.max_column

            # Add total row
            for col in range(1, max_col + 1):
                cell = ws.cell(row=max_row + 1, column=col)
                cell.font = total_style
                cell.fill = total_fill
                if col > 1:
                    cell.value = f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{max_row})"

            # Add total column
            for row in range(1, max_row + 2):
                cell = ws.cell(row=row, column=max_col + 1)
                cell.font = total_style
                cell.fill = total_fill
                if row == 1:
                    cell.value = "Total"
                else:
                    cell.value = f"=SUM(B{row}:{get_column_letter(max_col)}{row})"

            def add_total_label(sheet_name):
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    max_row = ws.max_row   # Identify the next row to add 'Total'
                    ws.cell(row=max_row, column=1, value="Total")  # Write 'Total' in the first column

            # Apply the function to the 'Weekly Forecast' and 'Monthly Forecast' sheets
            add_total_label('Weekly Forecast')
            add_total_label('Monthly Forecast')

            # Adjust the width of the new total column
            forecast_sheet_width = 18
            ws.column_dimensions[get_column_letter(max_col + 1)].width = forecast_sheet_width
            # Save the workbook with the sheets reordered

    # Adjust column widths for customer sheets ('BAM002' to 'BAM011')
    for sheet_code in prefix_to_sheet.values():
        adjust_column_widths(sheet_code, customer_sheet_width)

    # Add total row and column for 'Weekly Forecast' and 'Monthly Forecast' sheets
    add_totals('Weekly Forecast')
    add_totals('Monthly Forecast')

    # Adjust column widths for 'Weekly Forecast' and 'Monthly Forecast' sheets
    adjust_column_widths('Weekly Forecast', weekly_forecast_width)
    adjust_column_widths('Monthly Forecast', monthly_forecast_width)


    # Save the workbook with the sheets reordered
    wb.save(output_filepath)

    # Load the workbook and access the 'Monthly Forecast' sheet
    wb = load_workbook(output_filepath)
    ws = wb['Monthly Forecast']

    # Check the value in the 8th row of Column 1 ('Account')
    account_value = ws['A3'].value

    # If the value is 'BAM003', update the cell to 'A9', else update it to 'A8'
    if account_value == 'BAM003':
        ws['A9'] = 'BAM011'
        ws['A10'] = 'BAM018'
        print("Updated the value in the 9th row of Column 1 to 'BAM011'.")
    else:
        ws['A8'] = 'BAM011'
        ws['A9'] = 'BAM018'
        print("Updated the value in the 8th row of Column 1 to 'BAM011'.")

    # Save the workbook
    wb.save(output_filepath)

        # Load the workbook for reading and writing
    wb = load_workbook(output_filepath)

    # Ensure 'Picklist' sheet exists, if not, create it
    if 'Picklist' not in wb.sheetnames:
        wb.create_sheet('Picklist')
    ws_picklist = wb['Picklist']

    # Define the column headers for the 'Picklist' sheet, moving 'Required Day' to the new position
    column_headers = ['Account', 'Stock Code', 'Issue', 'Required Date', 'Required Day', 'Required QTY', 'Reference', 'Location', 'Message', 'Week']

    # Write the column headers to the 'Picklist' sheet
    ws_picklist.delete_rows(2, ws_picklist.max_row)  # Clear existing data
    for col, header in enumerate(column_headers, start=1):
        ws_picklist.cell(row=1, column=col, value=header)

    # Define the weeks of interest in the desired order
    weeks_of_interest = ['Week - 1', 'Week - 2', 'Week - 3', 'Week - 4', 'Week - 5', 'Week - 6', 'Week - 7']
    extended_weeks_of_interest = ['Arrears'] + [f'Week - {i}' for i in range(1, 8)]

    # Define the correct order of days
    days_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

    # Initialize a list to store rows to be added to the 'Picklist' sheet
    picklist_rows = []
    services_picklist_rows = []

    # Loop through each customer account sheet
    for sheet_name in prefix_to_sheet.values():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):  # Assuming row 1 has headers
                if sheet_name == 'BAM003' and 'BAM003' in wb.sheetnames and row[9] in extended_weeks_of_interest:
                    # Parse 'Required Date' and get 'Required Day'
                    try:
                        required_date = datetime.strptime(row[2], '%d/%m/%Y')
                        required_day = required_date.strftime('%A')
                    except ValueError:
                        required_day = 'Invalid Date'  # Handle invalid date format
                    
                    # Reorder row data to exclude 'Last Delivery Note' and 'Last Delivery Date'
                    new_row = [sheet_name] + list(row[:2]) + [row[2], required_day] + list(row[3:7]) + [row[9]]
                    services_picklist_rows.append(new_row)
                elif sheet_name != 'BAM003' and row[9] in weeks_of_interest:
                    # Parse 'Required Date' and get 'Required Day'
                    try:
                        required_date = datetime.strptime(row[2], '%d/%m/%Y')
                        required_day = required_date.strftime('%A')
                    except ValueError:
                        required_day = 'Invalid Date'  # Handle invalid date format
                    
                    # Reorder row data to exclude 'Last Delivery Note' and 'Last Delivery Date'
                    new_row = [sheet_name] + list(row[:2]) + [row[2], required_day] + list(row[3:7]) + [row[9]]
                    picklist_rows.append(new_row)

    # Sort the rows by 'Account' and then by 'Week'
    picklist_rows.sort(key=lambda x: (x[0], weeks_of_interest.index(x[-1]), days_order.index(x[4])))

    # Write the sorted rows to the 'Picklist' sheet
    for row_data in picklist_rows:
        ws_picklist.append(row_data)

    # Save the workbook with the new 'Picklist' sheet
    wb.save(output_filepath)

    print(f"Data for weeks {', '.join(weeks_of_interest)} gathered, sorted by account and week, and written to 'Picklist' sheet in {output_filepath}")

    # Load the workbook for reading and writing
    wb = load_workbook(output_filepath)

    # Access the 'Picklist' sheet
    ws_picklist = wb['Picklist']

    # Set the column width to 18 for all columns in 'Picklist'
    for col in ws_picklist.columns:
        ws_picklist.column_dimensions[get_column_letter(col[0].column)].width = 18

    # Find the last row and column with data to define the table range
    last_row = ws_picklist.max_row
    last_col = ws_picklist.max_column

    # Define the table range based on the data for 'Picklist'
    table_range_picklist = f"A1:{get_column_letter(last_col)}{last_row}"

    # Create a table for the range in 'Picklist'
    table_picklist = Table(displayName="PicklistTable", ref=table_range_picklist)

    # Add a default table style with striped rows and banded columns for 'Picklist'
    style_picklist = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table_picklist.tableStyleInfo = style_picklist

    # Add the table to the 'Picklist' worksheet
    ws_picklist.add_table(table_picklist)

    # Check if 'BAM003' sheet exists in the workbook
    if 'BAM003' in wb.sheetnames:
        # Ensure 'Services Picklist' sheet exists, if not, create it
        if 'Services Picklist' not in wb.sheetnames:
            wb.create_sheet('Services Picklist')
        ws_services_picklist = wb['Services Picklist']

        # Write the column headers to the 'Services Picklist' sheet
        ws_services_picklist.delete_rows(2, ws_services_picklist.max_row)  # Clear existing data
        for col, header in enumerate(column_headers, start=1):
            ws_services_picklist.cell(row=1, column=col, value=header)

        # Sort the rows by 'Account' and then by 'Week'
        services_picklist_rows.sort(key=lambda x: (x[0], extended_weeks_of_interest.index(x[-1]), days_order.index(x[4])))

        # Write the sorted rows to the 'Services Picklist' sheet
        for row_data in services_picklist_rows:
            ws_services_picklist.append(row_data)

        # Set the column width to 18 for all columns in 'Services Picklist'
        for col in ws_services_picklist.columns:
            ws_services_picklist.column_dimensions[get_column_letter(col[0].column)].width = 18

        # Find the last row and column with data to define the table range for 'Services Picklist'
        last_row_services_picklist = ws_services_picklist.max_row
        last_col_services_picklist = ws_services_picklist.max_column

        # Define the table range based on the data for 'Services Picklist'
        table_range_services_picklist = f"A1:{get_column_letter(last_col_services_picklist)}{last_row_services_picklist}"

        # Create a table for the range in 'Services Picklist'
        table_services_picklist = Table(displayName="ServicesPicklistTable", ref=table_range_services_picklist)

        # Add a default table style with striped rows and banded columns for 'Services Picklist'
        style_services_picklist = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table_services_picklist.tableStyleInfo = style_services_picklist

        # Add the table to the 'Services Picklist' worksheet
        ws_services_picklist.add_table(table_services_picklist)

        print(f"Table created for 'Services Picklist' sheet and column widths set to 18 in {output_filepath}")

    # Save the workbook with the new tables and adjusted column widths
    wb.save(output_filepath)

    print(f"Table created for 'Picklist' sheet and column widths set to 18 in {output_filepath}")

#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------#

# Section - 2: Generating the PDFs for Weekly and Monthly Sales Forecasts and Picklist for 4 - weeks. #

def get_user_input_for_services():
        user_input = input("Also Prepare the Services Picklist, Type Y for Yes and N for No")
        if user_input.lower() == 'y':
            account_name = 'BAM003'

            generate_services_picklist_pdf(excel_file_path, services_picklist_output_directory_path, logo_path, account_name)
        else:
            print("Skipping Services Picklist as Services HSE is not provided.")


def CreatePDFfilesFromWorkBook():

    # Generate the PDF file
    generate_separate_picklist_pdf(excel_file_path, full_picklist_pdf_output_directory_path, logo_path, separate_pdf_output_directory_path)

    # Call the function to generate the PDF for the 'Monthly Forecast' sheet
    generate_monthly_forecast_pdf(excel_file_path, forecast_output_directory_path, logo_path)

    generate_weekly_forecast_pdf(excel_file_path, forecast_output_directory_path, logo_path)

    combine_pdf_files(separate_pdf_output_directory_path, full_picklist_pdf_output_directory_path)

    get_user_input_for_services()

    
    


def main():
    createExcelWorkbookfromHSEFiles()
    CreatePDFfilesFromWorkBook()


# Example usage
if __name__ == "__main__":
    main()


        